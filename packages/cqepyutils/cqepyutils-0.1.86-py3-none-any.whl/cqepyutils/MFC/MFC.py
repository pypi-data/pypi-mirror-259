import os
import yaml
from openpyxl import Workbook


class MFC:
    def __init__(self, project_base_path):
        self.project_base_path = project_base_path

    def create_folders_and_files(self):
        # Create Mass_File_Compare parent folder
        parent_folder = os.path.join(self.project_base_path, 'Mass_File_Compare')
        os.makedirs(parent_folder, exist_ok=True)

        folders = ['config', 'data', 'keywords', 'results', 'scripts', 'testcases']

        # Create subfolders under Mass_File_Compare
        for folder in folders:
            folder_path = os.path.join(parent_folder, folder)
            os.makedirs(folder_path, exist_ok=True)

        # Create config.yaml under config folder
        config_path = os.path.join(parent_folder, 'config', 'config.yaml')
        with open(config_path, 'w') as config_file:
            # You can customize the content of the YAML file as needed
            yaml.dump({'key': 'value'}, config_file)

        # Create subfolders and files under data folder
        data_subfolders = ['db_queries', 'icd', 'test_inputs']
        for subfolder in data_subfolders:
            subfolder_path = os.path.join(parent_folder, 'data', subfolder)
            os.makedirs(subfolder_path, exist_ok=True)

        # Create MFC_Test_Data.xlsx under test_inputs folder using openpyxl
        test_data_path = os.path.join(parent_folder, 'data', 'test_inputs', 'MFC_Test_Data.xlsx')
        wb = Workbook()
        ws = wb.active

        # Add sample data to the Excel file (you can customize this part)
        ws['A1'] = 'Sample'
        ws['B1'] = 'Data'

        # Save the workbook
        wb.save(test_data_path)

        # Create keyword.resource file under keywords folder
        keyword_path = os.path.join(parent_folder, 'keywords', 'keyword.resource')
        with open(keyword_path, 'w'):
            pass  # You can add content to the file if needed

        # Create Trigger_Execution.bat under results folder
        bat_file_path = os.path.join(parent_folder, 'results', 'Trigger_Execution.bat')
        with open(bat_file_path, 'w') as bat_file:
            # Add the specified content to the batch file
            bat_file.write('python -m robot --reportitle "Pepper Fusion Test Automation Dashboard" '
                           '--logtitle "Pepper Fusion Test Automation Dashboard" -N "MFC Test Suite" -d . '
                           '..\\testcases\\MFC_Test_Suite.robot')

        # Create suite_common_library.py under scripts folder
        py_file_path = os.path.join(parent_folder, 'scripts', 'suite_common_library.py')
        with open(py_file_path, 'w') as py_file:
            # Add the specified content to the Python file
            py_file.write('from PepperFusion.CommonUtils.DBUtils import DBUtils\n'
                          'from robot.api import logger\n'
                          'from robot.api.deco import keyword')

        # Create MFC_Test_Suite.robot under testcases folder
        template_file_path = os.path.join(parent_folder, 'testcases', 'MFC_Test_Suite_Template.robot')
        robot_file_path = os.path.join(parent_folder, 'testcases', 'MFC_Test_Suite.robot')

        # Read the existing content from the template file
        with open(template_file_path, 'r') as template_file:
            existing_content = template_file.read()

        # Write the existing content to the new file
        with open(robot_file_path, 'w') as robot_file:
            robot_file.write(existing_content)

        print(f"Content from '{template_file_path}' has been written to '{robot_file_path}'.")


if __name__ == "__main__":
    # Specify the base path where you want to create the folders and files
    base_path = '/path/to/your/project'

    # Instantiate the MFC class and create folders and files
    mfc_instance = MFC(base_path)
    mfc_instance.create_folders_and_files()
