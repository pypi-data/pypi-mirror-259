# -*- coding:utf-8 -*-
"""
@Time        : 2023/11/5
@File        : touch_file.py
@Author      : lyz
@version     : python 3
@Description :
"""
import os
import zipfile
from bs4 import BeautifulSoup
import htmlmin
import xml.etree.ElementTree as ET
from PyUtilToolKit.log import Info
from PyUtilToolKit.pathSetting import ensure_path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def change_path(path):
    '''
    改变当前工作目录到上级或更上级目录中
    :param path: "." 表示当前目录，".." 表示父级目录 (../../../../path)
    :return:
    '''
    current_dir = os.getcwd()               # 1/获取当前工作目录
    # print("当前工作目录：", current_dir)
    os.chdir(path)                          # 2/改变当前工作目录到上级目录 ..
    current_dir = os.getcwd()               # 3/获取目标工作目录
    # print("目标工作目录：", current_dir)
    return current_dir

def touch_file(path_filename):
    '''
    使用相对路径创建文件
    :param path: 相对地址 （'../../../../appium_learning'）
    :param filename: 文件名.文件类型 （test.txt）
    '''
    relative_path = ensure_path(path_filename)
    # print("touch_file",relative_path)
    with open(relative_path, "w+", encoding="utf-8") as file:
        file.write("Hello, World!")
    # 查看文件是否存在
    if os.path.exists(relative_path):
        # Log().log("文件已生成")
        return relative_path
    else:
        raise Exception("文件未生成")
def del_file(path):
    """删除目录下的文件"""
    list_path = os.listdir(path)
    for i in list_path:
        c_path = os.path.join(path, i)
        if os.path.isdir(c_path):
            del_file(c_path)
        else:
            os.remove(c_path)

def touch_zip(input_file=ensure_path("\\yiXie_9_report\\report.html"),
              output_file=ensure_path("\\report.zip")):
    """创建压缩文件"""
    with open(input_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
        # 使用BeautifulSoup解析HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        # 重新编码压缩HTML
        minified_html = htmlmin.minify(str(soup), remove_comments=True, remove_empty_space=True)
        # 创建压缩文件并将压缩后的HTML写入其中
        with zipfile.ZipFile(output_file, 'w') as zipf:
            zipf.writestr('compressed.html', minified_html)
    Info.log('ZIP文件创建成功')

def touch_excel(input_file=ensure_path("\\yiXie_9_report\\result.xml"),
              output_file=ensure_path("\\yiXie_9_report\\result_report.xlsx")):
    '''
    创建excel格式的报告文件
    '''
    # 解析 XML 数据
    tree = ET.parse(input_file)
    root = tree.getroot()

    # 创建 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 写入表头
    headers = ['Serial Number', 'Class name', 'Test Name', 'Details', 'Status']
    ws.append(headers)

    # 提取<testcase><failure>内的数据
    num = 1
    for testcase in root.iter('testcase'):
        status = "成功"
        failure_tag = testcase.find('failure')
        if failure_tag is not None:
            status = "失败"
            classname = testcase.get('classname')
            testname = testcase.get('name')
            details = failure_tag.text
            ws.append([num, classname, testname, details, status])
            num += 1
        elif failure_tag is None:
            classname = testcase.get('classname')
            testname = testcase.get('name')
            ws.append([num, classname, testname, "AssertionCorrect", status])
            num += 1

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 筛选
    ws.auto_filter.ref = ws.dimensions

    # 设置所有字体为微软雅黑并将标题行加粗
    font = Font(name='Times New Roman')
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    # 标题行字体加粗
    for cell in ws[1]:
        cell.font = Font(bold=True, name='Times New Roman')

    # 冻结标题行
    ws.freeze_panes = 'A2'

    # 设置第4列宽度为x字符
    column_letter = get_column_letter(4)
    ws.column_dimensions[column_letter].width = 55

    # 标题行底色为黄色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = yellow_fill

    # 设置失败 色
    for row in ws.iter_rows():
        for cell in row:
            if '失败' in str(cell.value):
                # 获取当前单元格所在的行号和列号
                row_number = cell.row
                column_letter = cell.column_letter
                # 字体、行 上色
                for cell_in_row in ws[row_number]:
                    fill = PatternFill(start_color="FFF4EF", end_color="FFF4EF", fill_type="solid")
                    cell_in_row.fill = fill
                    font = Font(name='Times New Roman', color="CC2936", bold=True)
                    cell_in_row.font = font

                break

    # 初始化计数器
    fail_count = 0
    success_count = 0

    # 遍历表格中第5列的所有单元格，并统计包含的“失败”和“成功”数量
    for cell in ws['E']:
        if cell.value == '失败':
            fail_count += 1
        elif cell.value == '成功':
            success_count += 1

    # 将统计结果写入第7列第2行的格子中
    ws.cell(row=1, column=6, value=f"成功：{success_count}")
    ws.cell(row=1, column=7, value=f"失败：{fail_count}")

    '''新的工作表'''
    # 筛选失败数据
    filtered_data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[4] == "失败":
            filtered_data.append(row)

    # 将筛选结果写入新的工作表
    filtered_ws = wb.create_sheet(title="失败数据")
    filtered_ws.append(['Serial', 'Class name', 'Test Name', 'Details', 'Status'])
    for data in filtered_data:
        filtered_ws.append(data)

    # 设置所有字体为微软雅黑并将标题行加粗
    font = Font(name='Times New Roman')
    for row in filtered_ws.iter_rows():
        for cell in row:
            cell.font = font

    # 标题行字体加粗
    for cell in filtered_ws[1]:
        cell.font = Font(bold=True, name='Times New Roman')

    # 冻结标题行
    filtered_ws.freeze_panes = 'A2'

    # 筛选
    filtered_ws.auto_filter.ref = ws.dimensions

    # 标题行底色为黄色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in filtered_ws[1]:
        cell.fill = yellow_fill

    # 设置单行为X色，双行为X色
    for row_idx, row in enumerate(filtered_ws.iter_rows(min_row=2), start=1):
        if row_idx % 2 == 0:
            # 双数行
            for cell in row:
                cell.fill = PatternFill(start_color='FFFFDB', end_color='FFFFDB', fill_type='solid')

        if row_idx % 2 != 0:
            # 单数行
            for cell in row:
                cell.fill = PatternFill(start_color='F6F9F0', end_color='F6F9F0', fill_type='solid')

    # 自适应列宽
    for column in filtered_ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        # 设置列宽度
        column_letter = get_column_letter(2)
        filtered_ws.column_dimensions[column_letter].width = 25
        column_letter = get_column_letter(3)
        filtered_ws.column_dimensions[column_letter].width = 25
        column_letter = get_column_letter(4)
        filtered_ws.column_dimensions[column_letter].width = 157

    # 获取数据范围
    data_range = filtered_ws['A1':filtered_ws.cell(filtered_ws.max_row, filtered_ws.max_column).coordinate]
    # 将文字垂直居中
    for row in data_range:
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    # 将第一列单元格进行水平居中对齐
    for row in filtered_ws.iter_rows():
        cell = row[0]
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列自动换行
    for cell in filtered_ws["D"]:
        cell.alignment = Alignment(wrapText=True)

    # 插入空行
    for row_idx in range(10000):
        if row_idx % 2 == 1:
            filtered_ws.insert_rows(row_idx + 1)

    # 创建内部边框样式
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        diagonal=Side(border_style="thin")
    )

    # 设置所有单元格的内部边框
    for row in filtered_ws.iter_rows():
        for cell in row:
            cell.border = border

    # 检查每一行是否都没有数据，将底色设置为X色
    for row in filtered_ws.iter_rows():
        if not any(cell.value for cell in row):
            for cell in row:
                cell.fill = PatternFill(start_color='F6F9F0', end_color='F6F9F0', fill_type='solid')

    # 保存 Excel 文件
    wb.save(output_file)
    Info.log(f"成功将数据写入Excel文件：{output_file}")

if __name__ == '__main__':
    touch_excel()